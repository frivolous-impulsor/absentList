#By Mansor and olivger 

import tkinter as tk
import os
from os import listdir
from os.path import isfile, join, exists
import openpyxl
import re
import csv
import shutil
import openpyxl.worksheet


possibleColTitle = ["id", "first name", "name", "student id"]
searchField = ['id', 'student id']
myPath = os.getcwd()

def getCheckInLogDir():
    dirs = [o for o in os.listdir('.') if (os.path.isdir(o) and ("check" in o.lower() or "log" in o.lower()))]
    if len(dirs) == 1:
        return dirs[0]
    raise ValueError("only one subdirectory allowed. That being check in logs, being named \"checkInLogs\"")

logDir = join(myPath, getCheckInLogDir())

def getParameters():
    masterFile: str
    startTime: str
    endTime: str
    def getMasterFile():
        files = [f for f in os.listdir('.') if (os.path.isfile(f) and ".xlsx" in f)]
        if len(files) != 1:
            raise ValueError("place exactly one master file in xlsx format at root directory")
        masterFile = files[0]
        return masterFile

    def timeFormCheck(time):
        return re.match(r"\d+/\d+/\d+ \d+:\d+:\d+", time)
    
    window = tk.Tk()

    lblDate = tk.Label(text="Convocation Date: mm/dd/yyyy")
    entDate = tk.Entry()

    lblStart = tk.Label(text="start time: hh:mm")
    entStart = tk.Entry()
    lblEnd = tk.Label(text="end time: hh:mm")
    entEnd = tk.Entry()
    def runScript():
        nonlocal masterFile
        nonlocal startTime
        nonlocal endTime

        masterFile = getMasterFile()
        startTime = f"{entDate.get()} {entStart.get()}:00"
        endTime = f"{entDate.get()} {entEnd.get()}:00"
        if not (timeFormCheck(startTime) and timeFormCheck(endTime)):
            raise ValueError("time form not valid")
        window.destroy()

    btnConfirm = tk.Button(text="confirm", command=runScript)

    lblDate.pack()
    entDate.pack()
    lblStart.pack()
    entStart.pack()
    lblEnd.pack()
    entEnd.pack()
    btnConfirm.pack()
    window.mainloop()
    return (masterFile, startTime, endTime)

parameters = getParameters()

masterFile = parameters[0]
startTime = parameters[1]
endTime = parameters[2]

def translateCSV2XLSX(csvAddress: str) -> str:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    with open(csvAddress, encoding="utf8") as csvFile:
        csvData = csv.reader(csvFile, delimiter=',')
        for row in csvData:
            sheet.append(row)
    xlsxAddress = csvAddress.replace(".csv", ".xlsx")
    workbook.save(xlsxAddress)
    return xlsxAddress

def ensureXLSX(fileName) -> str:
    excelExt = '.xlsx'
    csvExt = '.csv'
    if excelExt in fileName and not (csvExt in fileName):
        return fileName
    if csvExt in fileName and not (excelExt in fileName):
        return translateCSV2XLSX(fileName)
    raise TypeError("checkInLogs consists files outside xlsx or csv, remove them")

def getLogFiles() -> list[str]:
    masterName = parameters
    logFiles = set()
    for fileName in listdir(logDir):
        if masterName == fileName: raise FileExistsError("master file should NOT exist in check in log directory, remove it")
        fileAddress = join(logDir, fileName)
        if isfile(fileAddress):
            xlsxAddress = ensureXLSX(fileAddress)
            logFiles.add(xlsxAddress)
            
        else:
            raise FileExistsError(f"non-file object in checkInLogs, remove {fileName}")
    return logFiles

def findTitleRow(sheet) -> int:
    for r in range(1, sheet.max_row+1):
        for c in range(1, sheet.max_column+1):
            if (sheet.cell(r, c).value != None) and type(sheet.cell(r,c).value) == type('str') and (sheet.cell(r, c).value.lower() in possibleColTitle):
                return r
    raise ValueError("No Proper Title Row Detected, Ensure The Tile Row Contains One Of The Possible Col Title")

def findColByTitles(sheet, titles: list[str]):
    titleRow = findTitleRow(sheet)
    for c in range(1, sheet.max_column+1):
        if (type(sheet.cell(titleRow, c).value) == type('str')) and (sheet.cell(titleRow, c).value.lower() in titles):
            return c
    raise ValueError("no colume with listed title found on title row")

def dateTimeStr2Tuple(dateTime: str):
    regex = r"(\d+)/(\d+)/(\d+) (\d+):(\d+):(\d+)"
    match = re.search(regex, dateTime)
    dateTimeTuple = tuple(match.group(i+1) for i in range(6))
    return dateTimeTuple



def findDataStartingRow(sheet, anchorDateTime:str) -> int:
    def removeLeadZero(str: str):
        str.lstrip('0')
        if str == '':
            str = '0'
        return str
    titleRow = findTitleRow(sheet)
    col = findColByTitles(sheet, ['time', 'timestamp'])
    maxRow = sheet.max_row
    anchorDateTime = dateTimeStr2Tuple(anchorDateTime)
    for row in range(titleRow+1, maxRow+1):
        currentDateTime = sheet.cell(row, col).value
        try: currentDateTime = dateTimeStr2Tuple(currentDateTime)
        except TypeError: continue
        if currentDateTime[:3] == anchorDateTime[:3]:
            hourI = 3
            minI = 4
            currentHour = int(removeLeadZero(currentDateTime[hourI]))
            anchorHour = int(removeLeadZero(anchorDateTime[hourI]))
            currentMin = int(removeLeadZero(currentDateTime[minI]))
            anchorMin = int(removeLeadZero(anchorDateTime[minI]))
            if (currentHour == anchorHour and currentMin >= anchorMin) or (currentHour > anchorHour):
                return row
    raise ValueError(f"No value start after {anchorDateTime}, reconfirm the excel and/or date time")

def findDataEndingRow(sheet, anchorDateTime: str) -> int:
    titleRow = findTitleRow(sheet)
    col = findColByTitles(sheet, ['time', 'timestamp'])
    maxRow = sheet.max_row
    endRow = 0
    anchorDateTime = dateTimeStr2Tuple(anchorDateTime)
    for row in range(titleRow+1, maxRow+1):
        currentDateTime = sheet.cell(row, col).value
        try: currentDateTime = dateTimeStr2Tuple(currentDateTime)
        except TypeError: continue
        if currentDateTime[:3] == anchorDateTime[:3]:
            if currentDateTime[3] < anchorDateTime[3] or (currentDateTime[3] == anchorDateTime[3] and currentDateTime[4] <= anchorDateTime[4]): 
                endRow = row
    if endRow == 0: raise ValueError("no entry in check log with given date")
    return endRow

def setIDsDict(path: str, idDict: dict, isLog: bool):
    book = openpyxl.load_workbook(path)
    sheetNames = book.sheetnames
    sheet = book[sheetNames[0]]
    idCol = findColByTitles(sheet, searchField)
    if isLog:
        try: startRow = findDataStartingRow(sheet, startTime)
        except ValueError: return
        endRow = findDataEndingRow(sheet, endTime)
    else:
        startRow = findTitleRow(sheet)
        endRow = sheet.max_row
    for r in range(startRow, endRow+1):
        id = str(sheet.cell(r, idCol).value)
        if id != str(None) and id.isnumeric():
            while len(id) < 9:
                id = '0'+id
            idDict[id] = isLog

def diffIDs(masterPath: str, minorPaths: list[str]) -> dict:
    IDdict = {}
    setIDsDict(masterPath, IDdict, False)
    for minor in minorPaths:
        setIDsDict(minor, IDdict, True)
    return IDdict

def createDir(path) -> int:
    if not exists(path):
        os.makedirs(path)
        return 1
    return 0

def writeList(idDict, resultDir: str) -> None:
    absentName = "absentList.xlsx"
    attendName = "attendList.xlsx"
    absentdst = join(resultDir, absentName)
    attenddst = join(resultDir, attendName)
    destAbsentAddress = shutil.copyfile(masterFile, absentdst)
    destAttendAddress = shutil.copyfile(masterFile, attenddst)
    
    absentBook = openpyxl.load_workbook(destAbsentAddress)
    absentSheet = absentBook[(absentBook.sheetnames)[0]]

    attendBook = openpyxl.load_workbook(destAttendAddress)
    attendSheet = attendBook[(attendBook.sheetnames)[0]]


    idCol = findColByTitles(absentSheet, searchField)
    rng = reversed(range(1, absentSheet.max_row+1))
    for row in rng:
        cellValue = absentSheet.cell(row, idCol).value
        if cellValue is not None and (cellValue in idDict.keys()):
            if idDict[cellValue]:
                absentSheet.delete_rows(row)
            else:
                attendSheet.delete_rows(row)
    
    absentBook.save(destAbsentAddress)
    attendBook.save(destAttendAddress)

def mergeCheckinLogs(logFileAddresses: list, dstDir: str) -> None:
    mergeName = "mergedCheckedIn.xlsx"
    mergeAddress = join(dstDir, mergeName)
    mergeBook = openpyxl.Workbook()
    mergeSheet = mergeBook.active

    titled = False
    titleFileAddress = ""
    for fileAddress in logFileAddresses:
        if not titled:
            titleFileAddress = fileAddress
            titled = True
        else:
            break

    titleBook = openpyxl.load_workbook(titleFileAddress)
    titleSheet = titleBook[(titleBook.sheetnames)[0]]
    titleRow = findTitleRow(titleSheet)
    titleVals = [cell.value for cell in titleSheet[titleRow]]
    mergeSheet.append(titleVals)

    for fileAddress in logFileAddresses:
        logBook = openpyxl.load_workbook(fileAddress)
        sheet = logBook[(logBook.sheetnames)[0]]
        idCod = findColByTitles(sheet, searchField)
        idIdx = idCod - 1
        try: startRow = findDataStartingRow(sheet, startTime)
        except ValueError: continue
        endRow = findDataEndingRow(sheet, endTime)
        for row in sheet.iter_rows(min_row=startRow, max_row=endRow):
            if row[idIdx].value != '' and row[idIdx].value != None:
                rowVals = [cell.value for cell in row]
                idVal = rowVals[idIdx]
                idVal = str(idVal)
                
                while len(idVal) < 9:
                    idVal = '0'+ idVal
                rowVals[idIdx] = idVal 
                mergeSheet.append(rowVals)
        
    mergeBook.save(mergeAddress)
        

def main() -> None:
    logFileAddresses = getLogFiles()
    idDict = diffIDs(masterFile, logFileAddresses)
    resultDirName = "result"
    createDir(resultDirName)
    writeList(idDict, resultDirName)
    mergeCheckinLogs(logFileAddresses, resultDirName)


main()